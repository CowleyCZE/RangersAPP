from sqlalchemy.orm import Session
from . import models, schemas
from passlib.context import CryptContext
from .data_extraction import extract_key_data_from_text

# --- Hesla a uživatelé ---
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_user_by_username(db: Session, username: str):
    return db.query(models.User).filter(models.User.username == username).first()

def create_user(db: Session, user: schemas.UserCreate):
    hashed_password = get_password_hash(user.password)
    db_user = models.User(
        username=user.username,
        email=user.email,
        hashed_password=hashed_password,
        role=user.role or "user"
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

from minio import Minio
import pytesseract
from PIL import Image
from io import BytesIO
import pdfplumber
import openpyxl
from datetime import datetime
import re
from typing import Dict, Any, List, Optional
import logging

logger = logging.getLogger(__name__)

def process_pdf_document(file_content: bytes) -> str:
    """Zpracování PDF dokumentu a extrakce textu"""
    text = ""
    try:
        with pdfplumber.open(BytesIO(file_content)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        logger.error(f"Chyba při zpracování PDF: {e}")
    return text

def process_xlsx_document(file_content: bytes) -> str:
    """Zpracování Excel dokumentu a extrakce textu"""
    text = []
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text.append(f"=== {sheet_name} ===")
            for row in ws.iter_rows():
                row_text = " ".join(str(cell.value) for cell in row if cell.value is not None)
                if row_text.strip():
                    text.append(row_text)
    except Exception as e:
        logger.error(f"Chyba při zpracování XLSX: {e}")
    return "\n".join(text)

def process_docx_document(file_content: bytes) -> str:
    """Zpracování Word dokumentu a extrakce textu"""
    try:
        from docx import Document
        doc = Document(BytesIO(file_content))
        text = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text.append(paragraph.text)
        return "\n".join(text)
    except ImportError:
        logger.warning("python-docx není nainstalován, DOCX soubory nebudou zpracovány")
        return ""
    except Exception as e:
        logger.error(f"Chyba při zpracování DOCX: {e}")
        return ""

def get_project(db: Session, project_id: int):
    return db.query(models.Project).filter(models.Project.id == project_id).first()

def get_projects(db: Session, skip: int = 0, limit: int = 100):
    return db.query(models.Project).offset(skip).limit(limit).all()

def create_project(db: Session, project: schemas.ProjectCreate):
    db_project = models.Project(**project.dict(), owner_id=1) # Hardcoded owner_id
    db.add(db_project)
    db.commit()
    db.refresh(db_project)
    return db_project

def update_project(db: Session, project_id: int, project: schemas.ProjectCreate):
    db_project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if db_project:
        db_project.name = project.name
        db_project.description = project.description
        db.commit()
        db.refresh(db_project)
    return db_project

def delete_project(db: Session, project_id: int):
    db_project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if db_project:
        db.delete(db_project)
        db.commit()
    return db_project

def get_documents(db: Session, project_id: int, category: str | None = None, skip: int = 0, limit: int = 100):
    query = db.query(models.Document).filter(models.Document.project_id == project_id)
    if category:
        query = query.filter(models.Document.category == category)
    return query.offset(skip).limit(limit).all()

def create_progress_log(db: Session, progress_log: schemas.ProgressLogCreate, project_id: int):
    db_progress_log = models.ProgressLog(**progress_log.dict(), project_id=project_id)
    db.add(db_progress_log)
    db.commit()
    db.refresh(db_progress_log)
    return db_progress_log

def update_progress_log(db: Session, progress_log_id: int, progress_log: schemas.ProgressLogCreate):
    db_progress_log = db.query(models.ProgressLog).filter(models.ProgressLog.id == progress_log_id).first()
    if db_progress_log:
        db_progress_log.date = progress_log.date
        db_progress_log.percentage_completed = progress_log.percentage_completed
        db_progress_log.notes = progress_log.notes
        db.commit()
        db.refresh(db_progress_log)
    return db_progress_log

def delete_progress_log(db: Session, progress_log_id: int):
    db_progress_log = db.query(models.ProgressLog).filter(models.ProgressLog.id == progress_log_id).first()
    if db_progress_log:
        db.delete(db_progress_log)
        db.commit()
    return db_progress_log

def get_progress_logs(db: Session, project_id: int, skip: int = 0, limit: int = 100):
    return db.query(models.ProgressLog).filter(models.ProgressLog.project_id == project_id).offset(skip).limit(limit).all()

def get_project_overall_progress(db: Session, project_id: int):
    progress_logs = db.query(models.ProgressLog).filter(models.ProgressLog.project_id == project_id).all()
    if not progress_logs:
        return 0
    total_percentage = sum([log.percentage_completed for log in progress_logs])
    return total_percentage / len(progress_logs)

def perform_ocr_on_document(db: Session, document_id: int, minio_client: Minio):
    """Rozšířená funkce pro OCR a extrakci dat z různých typů dokumentů"""
    db_document = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not db_document:
        return None, None

    bucket_name = "ranger-bucket"
    try:
        response = minio_client.get_object(bucket_name, db_document.filename)
        file_content = response.read()
        
        # Zpracování podle typu dokumentu
        filename_lower = db_document.filename.lower()
        
        if filename_lower.endswith('.pdf'):
            text = process_pdf_document(file_content)
        elif filename_lower.endswith(('.xlsx', '.xls')):
            text = process_xlsx_document(file_content)
        elif filename_lower.endswith('.docx'):
            text = process_docx_document(file_content)
        elif filename_lower.endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp')):
            # Pro obrázky použijeme OCR
            image = Image.open(BytesIO(file_content))
            text = pytesseract.image_to_string(image, lang='ces+eng')
        else:
            logger.warning(f"Nepodporovaný typ souboru: {db_document.filename}")
            return None, None
        
        if not text.strip():
            logger.warning(f"Žádný text nebyl extrahován ze souboru: {db_document.filename}")
            return "", {}
        
        # Pokročilá extrakce klíčových dat
        extracted_data = extract_key_data_from_text(text)
        
        logger.info(f"Úspěšně zpracován dokument {db_document.filename}, extrahováno {len(text)} znaků")
        return text, extracted_data
        
    except Exception as e:
        logger.error(f"Chyba při zpracování dokumentu {document_id}: {e}")
        return None, None
    finally:
        if 'response' in locals():
            response.close()
            response.release_conn()

def get_total_projects_count(db: Session):
    return db.query(models.Project).count()

def get_completed_projects_count(db: Session):
    completed_projects = db.query(models.Project).join(models.ProgressLog).filter(models.ProgressLog.percentage_completed == 100).distinct().count()
    return completed_projects

def get_average_overall_progress(db: Session):
    all_projects = db.query(models.Project).all()
    if not all_projects:
        return 0
    
    total_overall_progress = 0
    for project in all_projects:
        total_overall_progress += get_project_overall_progress(db, project.id)
    return total_overall_progress / len(all_projects)

# --- Detekce anomálií ve fotodokumentaci ---
import cv2
import numpy as np

def detect_anomaly_in_image(document_id: int, minio_client: Minio):
    """
    Skutečná logika detekce anomálií ve fotodokumentaci pomocí OpenCV.
    Detekuje základní typy anomálií: chybějící/nové objekty, poškození, barevné odchylky.
    """
    bucket_name = "ranger-bucket"
    from .models import SessionLocal, Document
    db = SessionLocal()
    
    try:
        db_document = db.query(Document).filter(Document.id == document_id).first()
        if not db_document:
            return {"anomaly_detected": False, "message": "Dokument nenalezen."}
        
        response = minio_client.get_object(bucket_name, db_document.filename)
        file_content = response.read()
        image = Image.open(BytesIO(file_content)).convert('RGB')
        img_np = np.array(image)
        img_cv = cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)

        # 1. Detekce rozmazání (ostrost)
        gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        laplacian_var = cv2.Laplacian(gray, cv2.CV_64F).var()
        blurry = laplacian_var < 50  # Prahová hodnota pro rozmazání

        # 2. Detekce dominantní barvy (např. příliš červené = rez)
        mean_color = cv2.mean(img_cv)
        red_ratio = mean_color[2] / (mean_color[0] + mean_color[1] + mean_color[2] + 1e-5)
        rusty = red_ratio > 0.5

        # 3. Detekce poškození (velké tmavé oblasti)
        _, thresh = cv2.threshold(gray, 40, 255, cv2.THRESH_BINARY_INV)
        dark_area = np.sum(thresh == 255) / (thresh.shape[0] * thresh.shape[1])
        damaged = dark_area > 0.15

        # 4. Detekce chybějících objektů pomocí jednoduché kontury
        contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        large_contours = [cnt for cnt in contours if cv2.contourArea(cnt) > 5000]
        missing_parts = len(large_contours) < 2  # Očekáváme aspoň 2 velké objekty

        # Výsledek
        anomaly_types = []
        if blurry:
            anomaly_types.append("Rozmazaný snímek")
        if rusty:
            anomaly_types.append("Podezření na rez (převaha červené)")
        if damaged:
            anomaly_types.append("Velké tmavé oblasti - možné poškození")
        if missing_parts:
            anomaly_types.append("Chybějící části/objekty")

        if anomaly_types:
            return {
                "anomaly_detected": True,
                "message": "Byly detekovány anomálie.",
                "details": ", ".join(anomaly_types),
                "analysis": {
                    "blur_score": laplacian_var,
                    "red_ratio": red_ratio,
                    "dark_area_ratio": dark_area,
                    "contour_count": len(large_contours)
                }
            }
        else:
            return {
                "anomaly_detected": False,
                "message": "Žádné zjevné anomálie nebyly detekovány.",
                "details": None,
                "analysis": {
                    "blur_score": laplacian_var,
                    "red_ratio": red_ratio,
                    "dark_area_ratio": dark_area,
                    "contour_count": len(large_contours)
                }
            }
    except Exception as e:
        logger.error(f"Chyba při detekci anomálií: {e}")
        return {"anomaly_detected": False, "message": f"Chyba při zpracování: {e}"}
    finally:
        db.close()
        if 'response' in locals():
            response.close()
            response.release_conn()